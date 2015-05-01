#-*-coding:utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

def CheckWs(wslist, wb):
    wsnames = wb.get_sheet_names()
    for i in wslist:
        if not i[0] in wsnames:
            ws = wb.create_sheet()
            ws.title = i[0]
            ws.append(['Title', 'PubDate', 'infohash'])

def ConvFeedlist(wb):
    ws = wb["Feedlist"]
    rowsNum = CheckrowsNum(ws)
    wslist = []
    for i in range(2, rowsNum):
        Date = ws.cell(row = i, column = 1).value
        Title = ws.cell(row = i, column = 2).value
        Group = ws.cell(row = i, column = 3).value
        Rss = ws.cell(row = i, column = 4).value
        wslist.append([Date + "," + Title + "," + Group, Rss])
    return wslist

def CheckrowsNum(ws):
    test = ws.cell(row = 1, column = 1).value
    i = 1
    while test <> None:
        i = i + 1
        test = ws.cell(row = i, column = 1).value
    return i - 1

def CheckcolumnsNum(ws):
    test = ws.cell(row = 1, column = 1).value
    i = 1
    while test <> None:
        i = i + 1
        test = ws.cell(row = 1, column = i).value
    return i - 1

def QueryItem(ws):
    rowsNum = CheckrowsNum(ws)
    if rowsNum > 1:
        Title = ws.cell(row = rowsNum, column = 1).value
        PubDate = ws.cell(row = rowsNum, column = 2).value
    else:
        Title = ''
        PubDate = ''
    return Title, PubDate

def updateBan(wb, infolist):
    ws = wb["Latest"]
    ws.append(infolist)

def appendEntry(wb, itemTitle, infolist):
    ws = wb[itemTitle]
    ws.append(infolist)
